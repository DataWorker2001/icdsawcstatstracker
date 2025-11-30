# app.py
import streamlit as st
import pandas as pd
import io
import re
import plotly.graph_objects as go
from datetime import datetime
import plotly.express as px  # Add this at the top
import calendar
# Initialize session state
if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None
if 'formatting_instructions' not in st.session_state:
    st.session_state.formatting_instructions = None
if 'awc_set' not in st.session_state:
    st.session_state.awc_set = None
if 'found_awcs' not in st.session_state:
    st.session_state.found_awcs = None
if 'file_process_errors' not in st.session_state:
    st.session_state.file_process_errors = None
if 'government_holidays' not in st.session_state:
    st.session_state.government_holidays = []
# For Excel cell formatting
from openpyxl.styles import PatternFill

st.set_page_config(
    page_title="AWC Monthly Data Processor, Tracker & Visualizer",
    page_icon="🤖",
    layout="wide"
)

st.title("🏥 ICDS AWC Monthly Status Tracker")

st.markdown("### Multi-file Merger & Compliance Monitor")
# st.markdown("""
# Upload up to 100 CSV/XLSX files whose filenames contain a date in the format `dd_mm_yyyy`
# (or `dd_mm_yyyy(1)` etc). Also upload a single CSV/XLSX file that lists AWC names/codes (one column).
# The app will filter rows across all uploaded daily files for AWCs present in that list, add `Date`
# and `Day` columns (parsed from filename), and produce a combined file grouped by AWC.
# """)
st.markdown("""
Upload up to 150 CSV/XLSX files whose filenames contain a date in the format `dd_mm_yyyy`
(or `dd_mm_yyyy(1)` etc). Also upload a single CSV/XLSX file that lists AWC names/codes (one column).
The app will filter rows across all uploaded daily files for AWCs present in that list, add `Date`
and `Day` columns (parsed from filename), and produce a combined file grouped by AWC.
""")

# Color Legend
st.markdown("---")
st.markdown("### 📊 Color Coding Legend")

col_legend1, col_legend2, col_legend3, col_legend4 = st.columns(4)

with col_legend1:
    st.markdown("""
    <div style="background-color: #FFFF00; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #E6E600;">
        <strong style="color: #000;">🟡 YELLOW</strong><br>
        <span style="font-size: 13px; color: #333;">AWC didn't open = 1<br>(Centre Closed)</span>
    </div>
    """, unsafe_allow_html=True)

with col_legend2:
    st.markdown("""
    <div style="background-color: #FFA500; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #E69500;">
        <strong style="color: #000;">🟠 ORANGE</strong><br>
        <span style="font-size: 13px; color: #333;">AWC open = 0 & Total HCM = 0<br>(No Hot Cooked Meal)</span>
    </div>
    """, unsafe_allow_html=True)

with col_legend3:
    st.markdown("""
    <div style="background-color: #FFC0CB; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #FFB0BB;">
        <strong style="color: #000;">🌸 PINK</strong><br>
        <span style="font-size: 13px; color: #333;">Snack ≠ 0 on Mon/Wed/Fri<br>(Snack on Non-Snack Day)</span>
    </div>
    """, unsafe_allow_html=True)

with col_legend4:
    st.markdown("""
    <div style="background-color: #FF033E; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #FFB7BE;">
        <strong style="color: #000;">🔴 RED</strong><br>
        <span style="font-size: 13px; color: #333;">Snack = 0 on Tue/Thu/Sat<br>(Missing Snack on Snack Day)</span>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

st.markdown("""
**Processing Rules:**
- **Rule 0:** All Sunday rows AND Government Holiday rows are automatically removed from the final output
- **Rule 1:** Yellow highlighting indicates the AWC was closed
- **Rule 2:** Orange highlighting indicates no hot cooked meals were provided despite AWC being open
- **Rule 3:** Pink highlighting indicates snacks were given on days when they shouldn't be (Mon/Wed/Fri)
- **Rule 4:** Red highlighting indicates missing snacks on designated snack days (Tue/Thu/Sat)

*Note: Original data values are preserved; only cell highlighting is applied in the XLSX output.*
""")

# --- Inputs -----------------------------------------------------------------------------
col1, col2 = st.columns([2, 1])

with col1:
    st.subheader("1) Upload daily files (up to 100)")
    # uploaded_files = st.file_uploader(
    #     "Choose CSV or XLSX files (multiple). Filenames must contain date like _dd_mm_yyyy",
    #     type=["csv", "xlsx"],
    #     accept_multiple_files=True,
    #     help="Examples: file_name_ABC_01_01_2025.csv or file_name_DEF_01_01_2025(1).xlsx"
    # )
    uploaded_files = st.file_uploader(
    "Choose CSV or XLSX files (multiple). Filenames must contain date like _dd_mm_yyyy",
    type=["csv", "xlsx"],
    accept_multiple_files=True,
    help="Examples: file_name_ABC_01_01_2025.csv or file_name_DEF_01_01_2025(1).xlsx. Max 150 files."
)

with col2:
    st.subheader("2) Upload AWC list file (single)")
    awc_file = st.file_uploader(
        "AWC list (csv or xlsx) — single file with one column containing AWC names/codes",
        type=["csv", "xlsx"],
        accept_multiple_files=False
    )

st.markdown("---")

# Options
st.sidebar.header("Options")
awc_column_hint = st.sidebar.text_input("If you know the AWC column header, enter it (optional)", "")
output_format = st.sidebar.selectbox("Output format", ["xlsx", "csv"])
sort_by_date = st.sidebar.checkbox("Sort each AWC's rows by Date ascending", value=True)
display_sample = st.sidebar.number_input("Show up to N matching rows in preview", min_value=0, max_value=1000, value=10)
st.sidebar.markdown("---")
st.sidebar.subheader("🏛️ Government Holidays")
st.sidebar.markdown("Select dates that should be excluded from performance evaluation:")

holiday_date = st.sidebar.date_input(
    "Select holiday date",
    value=None,
    help="Select a government holiday date"
)

col_h1, col_h2 = st.sidebar.columns(2)
with col_h1:
    if st.sidebar.button("➕ Add Holiday"):
        if holiday_date and holiday_date not in st.session_state.government_holidays:
            st.session_state.government_holidays.append(holiday_date)
            st.sidebar.success(f"Added: {holiday_date.strftime('%d-%b-%Y')}")

with col_h2:
    if st.sidebar.button("🗑️ Clear All"):
        st.session_state.government_holidays = []
        st.sidebar.success("Holidays cleared!")

    if st.session_state.government_holidays:
        st.sidebar.markdown("**Current Holidays:**")
        for idx, hol in enumerate(sorted(st.session_state.government_holidays)):
            col_h3, col_h4 = st.sidebar.columns([4, 1])
            with col_h3:
                st.sidebar.text(f"📅 {hol.strftime('%d-%b-%Y')}")
            with col_h4:
                if st.sidebar.button("❌", key=f"del_{idx}"):
                    st.session_state.government_holidays.remove(hol)
                    st.rerun()
# Helper functions ----------------------------------------------------------------------
DATE_REGEX = re.compile(r'(\d{1,2}_\d{1,2}_\d{4})')

def extract_date_from_filename(filename: str):
    """Find last occurrence of dd_mm_yyyy in filename and return datetime.date or None."""
    name = filename.split("/")[-1].split("\\")[-1]
    matches = DATE_REGEX.findall(name)
    if not matches:
        return None
    date_str = matches[-1]
    try:
        d = datetime.strptime(date_str, "%d_%m_%Y").date()
        return d
    except Exception:
        return None

def detect_awc_column(df: pd.DataFrame, hint: str = ""):
    """Try to detect AWC column in dataframe."""
    if hint:
        for col in df.columns:
            if col.strip().lower() == hint.strip().lower():
                return col
        for col in df.columns:
            if hint.strip().lower() in col.strip().lower():
                return col

    for col in df.columns:
        if 'awc' in col.strip().lower():
            return col

    for col in df.columns:
        if pd.api.types.is_object_dtype(df[col]) or not pd.api.types.is_numeric_dtype(df[col]):
            return col

    return df.columns[0]

def read_table_from_upload(uploaded_file, skip_header=False):
    """Read CSV or XLSX file from UploadedFile into pandas DataFrame."""
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)

    try:
        if skip_header:
            if name.endswith(".csv"):
                return pd.read_csv(uploaded_file, skiprows=8)
            else:
                return pd.read_excel(uploaded_file, sheet_name=0, skiprows=8)
        else:
            if name.endswith(".csv"):
                return pd.read_csv(uploaded_file)
            else:
                return pd.read_excel(uploaded_file, sheet_name=0)
    except Exception:
        uploaded_file.seek(0)
        if name.endswith(".csv"):
            return pd.read_csv(uploaded_file, encoding="utf-8", engine="python")
        raise

def ensure_columns_exist(df: pd.DataFrame, cols_with_defaults: dict):
    """Ensure columns exist in df; if missing create with default value."""
    for c, default in cols_with_defaults.items():
        if c not in df.columns:
            df[c] = default
    return df

def apply_business_rules_and_highlights(df: pd.DataFrame, government_holidays=None):
     
    """Apply rules and return modified dataframe and formatting instructions."""
    df2 = df.copy()
    if government_holidays is None:
        government_holidays = []

    if 'Date' in df2.columns:
        df2['Date'] = pd.to_datetime(df2['Date'], errors='coerce')
        df2['Day'] = df2['Date'].dt.day_name()
    else:
        df2['Date'] = pd.NaT
        df2['Day'] = ''

    # RULE 0: Remove Sundays
    # RULE 0: Remove Sundays AND Government Holidays
    df2 = df2[df2['Day'] != 'Sunday'].reset_index(drop=True)
    
    # Remove government holidays
    if government_holidays:
        holiday_dates = [pd.Timestamp(h) for h in government_holidays]
        df2 = df2[~df2['Date'].isin(holiday_dates)].reset_index(drop=True)

    df2 = ensure_columns_exist(df2, {
        "AWC didn't open": 0,
        "Total HCM Given": 0,
        "Morning Snack Given": 0
    })

    for c in ["AWC didn't open", "Total HCM Given", "Morning Snack Given"]:
        df2[c] = pd.to_numeric(df2[c], errors='coerce').fillna(0).astype(int)

    formats = []

    YELLOW = "FFFF00"
    ORANGE = "FFA500"
    PINK = "FFC0CB"
    RED = "FF033E"

    # for idx, row in df2.iterrows():
    #     day = str(row.get('Day', '')).strip()
    #     awc_didnt_open = int(row.get("AWC didn't open", 0))
    #     total_hcm = int(row.get('Total HCM Given', 0))
    #     snack = int(row.get('Morning Snack Given', 0))

    #     if awc_didnt_open == 1:
    #         formats.append((idx, "AWC didn't open", YELLOW))
        
    #     if awc_didnt_open == 0 and total_hcm == 0:
    #         formats.append((idx, "Total HCM Given", ORANGE))
        
    #     if awc_didnt_open == 0 and day in ["Monday", "Wednesday", "Friday"] and snack != 0:
    #         formats.append((idx, "Morning Snack Given", PINK))
        
    #     if awc_didnt_open == 0 and day in ["Tuesday", "Thursday", "Saturday"] and snack == 0:
    #         formats.append((idx, "Morning Snack Given", RED))
            
    #     # Calculate Attendance Percentage
    #     if 'Total Children Attended' in df2.columns and 'Total Children 3 yr to 6 yr' in df2.columns:
    #         df2['Total Children Attended'] = pd.to_numeric(df2['Total Children Attended'], errors='coerce').fillna(0).astype(int)
    #         df2['Total Children 3 yr to 6 yr'] = pd.to_numeric(df2['Total Children 3 yr to 6 yr'], errors='coerce').fillna(0).astype(int)
    # Calculate Attendance BEFORE the loop (moved outside)
    # if 'Total Children Attended' in df2.columns and 'Total Children 3 yr to 6 yr' in df2.columns:
    #     df2['Total Children Attended'] = pd.to_numeric(df2['Total Children Attended'], errors='coerce').fillna(0).astype(int)
    #     df2['Total Children 3 yr to 6 yr'] = pd.to_numeric(df2['Total Children 3 yr to 6 yr'], errors='coerce').fillna(0).astype(int)
        
    #     # Calculate attendance as "attended/total"
    #     df2['Attendance Marked'] = df2.apply(
    #         lambda row: f"{int(row['Total Children Attended'])}/{int(row['Total Children 3 yr to 6 yr'])}" 
    #         if row['Total Children 3 yr to 6 yr'] > 0 else "0/0",
    #         axis=1
    #     )
        
    #     # Calculate attendance percentage
    #     df2['Attendance %'] = df2.apply(
    #         lambda row: round((row['Total Children Attended'] / row['Total Children 3 yr to 6 yr'] * 100), 1) 
    #         if row['Total Children 3 yr to 6 yr'] > 0 else 0.0,
    #         axis=1
    #     )
    # else:
    #     df2['Attendance Marked'] = "N/A"
    #     df2['Attendance %'] = 0.0

    for idx, row in df2.iterrows():
        day = str(row.get('Day', '')).strip()
        awc_didnt_open = int(row.get("AWC didn't open", 0))
        total_hcm = int(row.get('Total HCM Given', 0))
        snack = int(row.get('Morning Snack Given', 0))

        if awc_didnt_open == 1:
            formats.append((idx, "AWC didn't open", YELLOW))
        
        if awc_didnt_open == 0 and total_hcm == 0:
            formats.append((idx, "Total HCM Given", ORANGE))
        
        if awc_didnt_open == 0 and day in ["Monday", "Wednesday", "Friday"] and snack != 0:
            formats.append((idx, "Morning Snack Given", PINK))
        
        if awc_didnt_open == 0 and day in ["Tuesday", "Thursday", "Saturday"] and snack == 0:
            formats.append((idx, "Morning Snack Given", RED))
            # Calculate attendance as "attended/total"
            df2['Attendance Marked'] = df2.apply(
                lambda row: f"{int(row['Total Children Attended'])}/{int(row['Total Children 3 yr to 6 yr'])}" 
                if row['Total Children 3 yr to 6 yr'] > 0 else "0/0",
                axis=1
            )
            
            # Calculate attendance percentage
            df2['Attendance %'] = df2.apply(
                lambda row: round((row['Total Children Attended'] / row['Total Children 3 yr to 6 yr'] * 100), 1) 
                if row['Total Children 3 yr to 6 yr'] > 0 else 0.0,
                axis=1
            )
        else:
            df2['Attendance Marked'] = "N/A"
            df2['Attendance %'] = 0.0

    return df2, formats

# Processing ---------------------------------------------------------------------------
if st.button("Process files"):
    if not uploaded_files or len(uploaded_files) == 0:
        st.error("Please upload at least one daily file.")
    elif not awc_file:
        st.error("Please upload the AWC list file.")
    else:
        st.info(f"Processing {len(uploaded_files)} files...")
        
        try:
            awc_df = read_table_from_upload(awc_file, skip_header=False)
        except Exception as e:
            st.exception(f"Failed to read AWC file: {e}")
            st.stop()

        try:
            awc_col_list = detect_awc_column(awc_df, awc_column_hint)
            awc_series = awc_df[awc_col_list].dropna().astype(str).str.strip()
            awc_set = set(awc_series.unique())
            if len(awc_set) == 0:
                st.warning("AWC file read successfully but no values detected.")
        except Exception as e:
            st.exception(f"Error detecting AWC column: {e}")
            st.stop()

        combined_rows = []
        found_awcs = set()
        file_process_errors = []

        for f in uploaded_files[:150]:
            fname = f.name
            try:
                df = read_table_from_upload(f, skip_header=True)
            except Exception as e:
                file_process_errors.append((fname, str(e)))
                continue

            if df.shape[0] == 0:
                continue

            try:
                awc_col = detect_awc_column(df, awc_column_hint)
            except Exception:
                awc_col = df.columns[0]

            df[awc_col] = df[awc_col].astype(str).str.strip()

            matched = df[df[awc_col].isin(awc_set)].copy()
            found_awcs.update(matched[awc_col].unique())

            date_val = extract_date_from_filename(fname)
            if date_val is None:
                matched['__filename_date_missing'] = True
            else:
                matched['Date'] = pd.to_datetime(date_val)

            if 'Date' in matched.columns:
                matched['Day'] = matched['Date'].dt.day_name()
            else:
                matched['Day'] = ""

            matched['Source_Filename'] = fname

            if not matched.empty:
                combined_rows.append(matched)

        if combined_rows:
            combined_df = pd.concat(combined_rows, ignore_index=True, sort=False)
        else:
            combined_df = pd.DataFrame()

        if combined_df.empty:
            st.warning("No rows matched the AWCs from the AWC list.")
            if file_process_errors:
                st.write("Some files could not be processed:")
                for fname, err in file_process_errors:
                    st.write(f"- {fname}: {err}")
            st.stop()

        if 'Date' in combined_df.columns:
            combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce')
        else:
            combined_df['Date'] = pd.NaT
        combined_df['Day'] = combined_df.get('Day', '').fillna('')

        detected_awc_col = None
        for col in combined_df.columns:
            try:
                vals = combined_df[col].dropna().astype(str)
                inter = set(vals.unique()).intersection(awc_set)
                if len(inter) > 0:
                    detected_awc_col = col
                    break
            except Exception:
                continue
        if detected_awc_col is None:
            detected_awc_col = combined_df.columns[0]
        if detected_awc_col != 'AWC':
            combined_df = combined_df.rename(columns={detected_awc_col: 'AWC'})

        if sort_by_date:
            combined_df = combined_df.sort_values(by=['AWC', 'Date']).reset_index(drop=True)
        else:
            combined_df = combined_df.sort_values(by=['AWC']).reset_index(drop=True)

        rest_cols = [c for c in combined_df.columns if c not in ['AWC', 'Date', 'Day', 'Source_Filename']]
        final_cols = ['AWC', 'Date', 'Day', 'Source_Filename'] + rest_cols
        combined_df = combined_df[final_cols]

        # finalized_df, formatting_instructions = apply_business_rules_and_highlights(combined_df)
        finalized_df, formatting_instructions = apply_business_rules_and_highlights(
            combined_df, 
            government_holidays=st.session_state.government_holidays
        )

        # Store in session state to prevent vanishing on download
        st.session_state.processed_data = finalized_df
        st.session_state.formatting_instructions = formatting_instructions
        st.session_state.awc_set = awc_set
        st.session_state.found_awcs = found_awcs
        st.session_state.file_process_errors = file_process_errors
        
        # Display results from session state (prevents vanishing on download)
        # Display holiday exclusion info
    if st.session_state.government_holidays:
        st.info(f"🏛️ **{len(st.session_state.government_holidays)} Government Holiday(s) excluded from analysis:** " + 
                ", ".join([h.strftime('%d-%b-%Y') for h in sorted(st.session_state.government_holidays)]))
    if st.session_state.processed_data is not None:
        finalized_df = st.session_state.processed_data
        formatting_instructions = st.session_state.formatting_instructions
        awc_set = st.session_state.awc_set
        found_awcs = st.session_state.found_awcs
        file_process_errors = st.session_state.file_process_errors
    
    # Move all display code here
        
        st.success("✅ Processing finished and business rules applied.")
        
        col_sum1, col_sum2, col_sum3, col_sum4 = st.columns(4)
        
        yellow_count = sum(1 for (_, col, color) in formatting_instructions if color == "FFFF00")
        orange_count = sum(1 for (_, col, color) in formatting_instructions if color == "FFA500")
        pink_count = sum(1 for (_, col, color) in formatting_instructions if color == "FFC0CB")
        red_count = sum(1 for (_, col, color) in formatting_instructions if color == "FF033E")
        
        with col_sum1:
            st.markdown(f"""
            <div style="background-color: #FFFF00; padding: 10px; border-radius: 5px; text-align: center;">
                <strong style="color: #000;">🟡 AWC Closed</strong><br>
                <span style="font-size: 24px; color: #000;"><strong>{yellow_count}</strong></span>
            </div>
            """, unsafe_allow_html=True)
        
        with col_sum2:
            st.markdown(f"""
            <div style="background-color: #FFA500; padding: 10px; border-radius: 5px; text-align: center;">
                <strong style="color: #000;">🟠 No HCM</strong><br>
                <span style="font-size: 24px; color: #000;"><strong>{orange_count}</strong></span>
            </div>
            """, unsafe_allow_html=True)
        
        with col_sum3:
            st.markdown(f"""
            <div style="background-color: #FFC0CB; padding: 10px; border-radius: 5px; text-align: center;">
                <strong style="color: #000;">🌸 Extra Snack</strong><br>
                <span style="font-size: 24px; color: #000;"><strong>{pink_count}</strong></span>
            </div>
            """, unsafe_allow_html=True)
        
        with col_sum4:
            st.markdown(f"""
            <div style="background-color: #FF033E; padding: 10px; border-radius: 5px; text-align: center;">
                <strong style="color: #000;">🔴 Missing Snack</strong><br>
                <span style="font-size: 24px; color: #000;"><strong>{red_count}</strong></span>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        st.write(f"Files processed: {len(uploaded_files)} (limited to first 150).")
        st.write(f"Total matched rows (after removing Sundays): {len(finalized_df)}")
        st.write(f"Number of unique AWCs in AWC list: {len(awc_set)}")
        st.write(f"Number of AWCs found across uploaded files: {len(found_awcs)}")
        not_found = set(awc_set) - set(found_awcs)
        st.write(f"AWCs from list not present in any file: {len(not_found)}")
        if len(not_found) <= 50:
            st.write(sorted(list(not_found)))
        else:
            st.write(f"(showing first 50) {sorted(list(not_found))[:50]}")

        if file_process_errors:
            st.warning("Some files failed to be read:")
            for fname, err in file_process_errors:
                st.write(f"- {fname}: {err}")

        st.subheader("📊 Data Preview")
        st.markdown("*Preview of finalized combined results (first rows)*")
        if display_sample > 0:
            st.dataframe(finalized_df.head(display_sample), use_container_width=True)
        else:
            st.dataframe(finalized_df, use_container_width=True)
        
        st.markdown("---")
        
        # ANALYTICS SECTION
        st.subheader("📈 Detailed Analytics & Insights")
        
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Overview", "📅 Monthly Analysis", "🎯 Date-wise Deep Dive", "🏢 AWC-wise Analysis", "⚠️ Compliance Issues"])
        
        with tab1:
            st.markdown("#### Overall Statistics")
            
            col_a1, col_a2, col_a3 = st.columns(3)
            
            with col_a1:
                st.metric("Total Records Processed", len(finalized_df))
                st.metric("Unique AWCs", finalized_df['AWC'].nunique())
                st.metric("Date Range", f"{finalized_df['Date'].min().strftime('%d-%b-%Y') if pd.notna(finalized_df['Date'].min()) else 'N/A'} to {finalized_df['Date'].max().strftime('%d-%b-%Y') if pd.notna(finalized_df['Date'].max()) else 'N/A'}")
            
            with col_a2:
                total_closures = (finalized_df["AWC didn't open"] == 1).sum()
                total_no_hcm = ((finalized_df["AWC didn't open"] == 0) & (finalized_df["Total HCM Given"] == 0)).sum()
                st.metric("🟡 Total AWC Closures", total_closures)
                st.metric("🟠 Days without HCM (while open)", total_no_hcm)
            
            with col_a3:
                mon_wed_fri = finalized_df['Day'].isin(['Monday', 'Wednesday', 'Friday'])
                tue_thu_sat = finalized_df['Day'].isin(['Tuesday', 'Thursday', 'Saturday'])
                
                extra_snacks = ((finalized_df["AWC didn't open"] == 0) & 
                               (finalized_df["Morning Snack Given"] != 0) & 
                               mon_wed_fri).sum()
                missing_snacks = ((finalized_df["AWC didn't open"] == 0) & 
                                 (finalized_df["Morning Snack Given"] == 0) & 
                                 tue_thu_sat).sum()
                
                st.metric("🌸 Extra Snacks Given", extra_snacks)
                st.metric("🔴 Missing Required Snacks", missing_snacks)
        
            
            st.markdown("---")
            st.markdown("#### 📊 Compliance Rate")
            
            total_records = len(finalized_df)
            total_issues = yellow_count + orange_count + pink_count + red_count
            compliance_rate = ((total_records - total_issues) / total_records * 100) if total_records > 0 else 0
            
            col_comp1, col_comp2 = st.columns(2)
            with col_comp1:
                st.markdown(f"""
                <div style="background-color: {'#90EE90' if compliance_rate >= 80 else '#FFB6C1'}; 
                            padding: 20px; border-radius: 10px; text-align: center;">
                    <h2 style="color: #000; margin: 0;">{compliance_rate:.1f}%</h2>
                    <p style="color: #333; margin: 5px 0;">Overall Compliance Rate</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col_comp2:
                st.markdown(f"""
                <div style="background-color: #F0F0F0; padding: 20px; border-radius: 10px; text-align: center;">
                    <h2 style="color: #000; margin: 0;">{total_issues}</h2>
                    <p style="color: #333; margin: 5px 0;">Total Issues Detected</p>
                </div>
                """, unsafe_allow_html=True)
            # Attendance Statistics
            if 'Total Children Attended' in finalized_df.columns and 'Total Children 3 yr to 6 yr' in finalized_df.columns:
                st.markdown("---")
                st.markdown("#### 👥 Overall Attendance Statistics")
                
                total_attended_overall = finalized_df['Total Children Attended'].sum()
                total_children_overall = finalized_df['Total Children 3 yr to 6 yr'].sum()
                overall_attendance_rate = (total_attended_overall / total_children_overall * 100) if total_children_overall > 0 else 0
                
                col_att_ov1, col_att_ov2, col_att_ov3 = st.columns(3)
                
                with col_att_ov1:
                    st.metric("Total Children Registered (3-6 yr)", f"{total_children_overall:,}")
                    
                with col_att_ov2:
                    st.metric("Total Attendance Marked", f"{total_attended_overall:,}")
                
                with col_att_ov3:
                    st.metric("Average Attendance Rate", f"{overall_attendance_rate:.1f}%")
        
        with tab2:
            st.markdown("#### 📅 Monthly Distribution")
            
            day_summary = finalized_df.groupby('Day').agg({
                'AWC': 'count',
                "AWC didn't open": 'sum',
                "Total HCM Given": lambda x: (x == 0).sum(),
                "Morning Snack Given": 'sum'
            }).reset_index()
            day_summary.columns = ['Day', 'Total Records', 'Closures', 'No HCM Days', 'Total Snacks Given']
            
            day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
            day_summary['Day'] = pd.Categorical(day_summary['Day'], categories=day_order, ordered=True)
            day_summary = day_summary.sort_values('Day')
            
            st.dataframe(day_summary, use_container_width=True, hide_index=True)
            
            st.markdown("---")
            st.markdown("#### 📊 Visual Distribution by Day")
            
            col_vis1, col_vis2 = st.columns(2)
            
            with col_vis1:
                st.markdown("**Records by Day**")
                st.bar_chart(day_summary.set_index('Day')['Total Records'])
            
            with col_vis2:
                st.markdown("**Closures by Day**")
                st.bar_chart(day_summary.set_index('Day')['Closures'])
        
        with tab3:
            st.markdown("#### 🎯 Date-wise Deep Dive Analysis")
            
            available_dates = sorted(finalized_df['Date'].dropna().dt.date.unique())
            
            if len(available_dates) > 0:
                selected_date = st.selectbox(
                    "📅 Select a date to analyze:",
                    options=available_dates,
                    format_func=lambda x: x.strftime('%d %B %Y (%A)')
                )
                
                date_data = finalized_df[finalized_df['Date'].dt.date == selected_date].copy()
                
                if not date_data.empty:
                    day_name = date_data['Day'].iloc[0]
                    
                    st.markdown(f"### Analysis for {selected_date.strftime('%d %B %Y')} ({day_name})")
                    st.markdown("---")
                    
                    col_d1, col_d2, col_d3, col_d4 = st.columns(4)
                    
                    total_awcs = len(date_data)
                    closed_awcs = (date_data["AWC didn't open"] == 1).sum()
                    open_awcs = total_awcs - closed_awcs
                    no_hcm_awcs = ((date_data["AWC didn't open"] == 0) & (date_data["Total HCM Given"] == 0)).sum()
                    
                    with col_d1:
                        st.markdown(f"""
                        <div style="background-color: #E3F2FD; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #2196F3;">
                            <h3 style="color: #1976D2; margin: 0;">{open_awcs}/{total_awcs}</h3>
                            <p style="color: #333; margin: 5px 0; font-size: 14px;">AWCs Opened</p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col_d2:
                        st.markdown(f"""
                        <div style="background-color: #FFF9C4; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #FBC02D;">
                            <h3 style="color: #F57F17; margin: 0;">{closed_awcs}</h3>
                            <p style="color: #333; margin: 5px 0; font-size: 14px;">🟡 Closures</p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col_d3:
                        st.markdown(f"""
                        <div style="background-color: #FFE0B2; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #FB8C00;">
                            <h3 style="color: #E65100; margin: 0;">{no_hcm_awcs}</h3>
                            <p style="color: #333; margin: 5px 0; font-size: 14px;">🟠 No HCM</p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    if day_name in ['Tuesday', 'Thursday', 'Saturday']:
                        snack_metric = ((date_data["AWC didn't open"] == 0) & (date_data["Morning Snack Given"] == 0)).sum()
                        correct_snacks = ((date_data["AWC didn't open"] == 0) & (date_data["Morning Snack Given"] != 0)).sum()
                        metric_label = "🔴 Missing Snacks"
                        metric_color = "#FFCDD2"
                        metric_border = "#E57373"
                    else:
                        snack_metric = ((date_data["AWC didn't open"] == 0) & (date_data["Morning Snack Given"] != 0)).sum()
                        correct_snacks = ((date_data["AWC didn't open"] == 0) & (date_data["Morning Snack Given"] == 0)).sum()
                        metric_label = "🌸 Extra Snacks"
                        metric_color = "#F8BBD0"
                        metric_border = "#EC407A"
                    
                    with col_d4:
                        st.markdown(f"""
                        <div style="background-color: {metric_color}; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid {metric_border};">
                            <h3 style="color: #000; margin: 0;">{snack_metric}</h3>
                            <p style="color: #333; margin: 5px 0; font-size: 14px;">{metric_label}</p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    st.markdown("---")
                    st.markdown("#### 📊 Distribution Charts")
                    
                    col_chart1, col_chart2 = st.columns(2)
                    
                    with col_chart1:
                        st.markdown("**AWC Opening Status**")
                        fig1 = go.Figure(data=[go.Pie(
                            labels=['Opened', 'Closed'],
                            values=[open_awcs, closed_awcs],
                            hole=0.4,
                            marker_colors=['#4CAF50', '#FFEB3B'],
                            textinfo='label+value+percent',
                            textfont_size=14
                        )])
                        fig1.update_layout(showlegend=True, height=350, margin=dict(l=20, r=20, t=30, b=20))
                        st.plotly_chart(fig1, use_container_width=True)
# Similarly replace the snack chart (around line 556-566) and HCM charts (around lines 578-596) with donut chart code I provided earlier.

# These 5 changes will fix both the vanishing issue and add donut charts!Claude can make mistakes. Please double-check responses. Sonnet 4.5
                    
                    with col_chart2:
                        if day_name in ['Tuesday', 'Thursday', 'Saturday']:
                            st.markdown("**Snack Distribution (Snack Day)**")
                            fig2 = go.Figure(data=[go.Pie(
                                labels=['Correct (Snack Given)', 'Missing Snack'],
                                values=[correct_snacks, snack_metric],
                                hole=0.4,
                                marker_colors=['#4CAF50', '#F44336'],
                                textinfo='label+value+percent',
                                textfont_size=14
                            )])
                        else:
                            st.markdown("**Snack Distribution (Non-Snack Day)**")
                            fig2 = go.Figure(data=[go.Pie(
                                labels=['Correct (No Snack)', 'Extra Snack Given'],
                                values=[correct_snacks, snack_metric],
                                hole=0.4,
                                marker_colors=['#4CAF50', '#FFC0CB'],
                                textinfo='label+value+percent',
                                textfont_size=14
                            )])
                        
                        fig2.update_layout(showlegend=True, height=350, margin=dict(l=20, r=20, t=30, b=20))
                        st.plotly_chart(fig2, use_container_width=True)
                        
                    st.markdown("---")
                    
                    col_hcm1, col_hcm2 = st.columns(2)
                    
                    with col_hcm1:
                        st.markdown("**Hot Cooked Meal Distribution**")
                        hcm_correct = ((date_data["AWC didn't open"] == 0) & (date_data["Total HCM Given"] != 0)).sum()
                        
                        fig3 = go.Figure(data=[go.Pie(
                            labels=['HCM Provided', 'No HCM (Open AWC)'],
                            values=[hcm_correct, no_hcm_awcs],
                            hole=0.4,
                            marker_colors=['#4CAF50', '#FF9800'],
                            textinfo='label+value+percent',
                            textfont_size=14
                        )])
                        fig3.update_layout(showlegend=True, height=350, margin=dict(l=20, r=20, t=30, b=20))
                        st.plotly_chart(fig3, use_container_width=True)
                    
                    with col_hcm2:
                        st.markdown("**Overall Compliance for This Date**")
                        date_issues = closed_awcs + no_hcm_awcs + snack_metric
                        date_compliance = ((total_awcs - date_issues) / total_awcs * 100) if total_awcs > 0 else 0
                        
                        fig4 = go.Figure(data=[go.Pie(
                            labels=['Compliant', 'Non-Compliant'],
                            values=[total_awcs - date_issues, date_issues],
                            hole=0.4,
                            marker_colors=['#4CAF50', '#F44336'],
                            textinfo='label+value+percent',
                            textfont_size=14
                        )])
                        fig4.update_layout(showlegend=True, height=350, margin=dict(l=20, r=20, t=30, b=20))
                        st.plotly_chart(fig4, use_container_width=True)
                        
                        st.metric("Compliance Rate", f"{date_compliance:.1f}%")
                    # Calculate attendance metrics for this date
                    if 'Total Children Attended' in date_data.columns and 'Total Children 3 yr to 6 yr' in date_data.columns:
                        total_attended = date_data['Total Children Attended'].sum()
                        total_children = date_data['Total Children 3 yr to 6 yr'].sum()
                        attendance_rate = (total_attended / total_children * 100) if total_children > 0 else 0
                        
                        st.markdown("---")
                        st.markdown("#### 👥 Attendance Overview")
                        
                        col_att1, col_att2, col_att3 = st.columns(3)
                        
                        with col_att1:
                            st.markdown(f"""
                            <div style="background-color: #E8F5E9; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #4CAF50;">
                                <h3 style="color: #2E7D32; margin: 0;">{total_attended}/{total_children}</h3>
                                <p style="color: #333; margin: 5px 0; font-size: 14px;">Total Attendance</p>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col_att2:
                            st.markdown(f"""
                            <div style="background-color: #FFF3E0; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #FF9800;">
                                <h3 style="color: #E65100; margin: 0;">{attendance_rate:.1f}%</h3>
                                <p style="color: #333; margin: 5px 0; font-size: 14px;">Attendance Rate</p>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col_att3:
                            absent_children = total_children - total_attended
                            st.markdown(f"""
                            <div style="background-color: #FFEBEE; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #F44336;">
                                <h3 style="color: #C62828; margin: 0;">{absent_children}</h3>
                                <p style="color: #333; margin: 5px 0; font-size: 14px;">Absent Children</p>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        st.markdown("---")
                        col_att_chart1, col_att_chart2 = st.columns(2)
                        
                        with col_att_chart1:
                            st.markdown("**Attendance Distribution**")
                            
                            fig_att = go.Figure(data=[go.Pie(
                                labels=['Present', 'Absent'],
                                values=[total_attended, absent_children],
                                hole=0.4,
                                marker_colors=['#4CAF50', '#F44336'],
                                textinfo='label+value+percent',
                                textfont_size=14
                            )])
                            fig_att.update_layout(showlegend=True, height=350, margin=dict(l=20, r=20, t=30, b=20))
                            st.plotly_chart(fig_att, use_container_width=True)
                        
                        with col_att_chart2:
                            st.markdown("**AWC-wise Attendance Performance**")
                            
                            # Calculate attendance rate per AWC
                            awc_attendance = date_data.groupby('AWC').agg({
                                'Total Children Attended': 'sum',
                                'Total Children 3 yr to 6 yr': 'sum'
                            }).reset_index()
                            
                            awc_attendance['Attendance %'] = (
                                awc_attendance['Total Children Attended'] / 
                                awc_attendance['Total Children 3 yr to 6 yr'] * 100
                            ).fillna(0).round(1)
                            
                            # Categorize AWCs by attendance
                            excellent = (awc_attendance['Attendance %'] >= 90).sum()
                            good = ((awc_attendance['Attendance %'] >= 75) & (awc_attendance['Attendance %'] < 90)).sum()
                            needs_improvement = (awc_attendance['Attendance %'] < 75).sum()
                            
                            fig_att2 = go.Figure(data=[go.Pie(
                                labels=['Excellent (≥90%)', 'Good (75-89%)', 'Needs Improvement (<75%)'],
                                values=[excellent, good, needs_improvement],
                                hole=0.4,
                                marker_colors=['#4CAF50', '#FF9800', '#F44336'],
                                textinfo='label+value+percent',
                                textfont_size=12
                            )])
                            fig_att2.update_layout(showlegend=True, height=350, margin=dict(l=20, r=20, t=30, b=20))
                            
                            # Make chart interactive - capture click event
                            selected_category = st.plotly_chart(fig_att2, use_container_width=True, on_select="rerun", selection_mode="points", key="attendance_chart")
                            
                            # Show AWCs based on selection or all categories by default
                            st.markdown("---")
                            st.markdown("#### 📋 AWC Attendance Details by Category")

                            # Create tabs for each category
                            att_tab1, att_tab2, att_tab3 = st.tabs([
                                f"🟢 Excellent (≥90%) - {excellent} AWCs", 
                                f"🟠 Good (75-89%) - {good} AWCs", 
                                f"🔴 Needs Improvement (<75%) - {needs_improvement} AWCs"
                            ])

                            with att_tab1:
                                excellent_awcs = awc_attendance[awc_attendance['Attendance %'] >= 90].sort_values('Attendance %', ascending=False)
                                if not excellent_awcs.empty:
                                    excellent_display = excellent_awcs.copy()
                                    excellent_display['Attendance'] = excellent_display.apply(
                                        lambda x: f"{int(x['Total Children Attended'])}/{int(x['Total Children 3 yr to 6 yr'])} ({x['Attendance %']:.1f}%)",
                                        axis=1
                                    )
                                    st.dataframe(
                                        excellent_display[['AWC', 'Attendance', 'Attendance %']],
                                        use_container_width=True,
                                        hide_index=True
                                    )
                                    st.success(f"✅ {len(excellent_awcs)} AWCs maintaining excellent attendance!")
                                else:
                                    st.info("No AWCs in this category.")

                            with att_tab2:
                                good_awcs = awc_attendance[
                                    (awc_attendance['Attendance %'] >= 75) & 
                                    (awc_attendance['Attendance %'] < 90)
                                ].sort_values('Attendance %', ascending=False)
                                if not good_awcs.empty:
                                    good_display = good_awcs.copy()
                                    good_display['Attendance'] = good_display.apply(
                                        lambda x: f"{int(x['Total Children Attended'])}/{int(x['Total Children 3 yr to 6 yr'])} ({x['Attendance %']:.1f}%)",
                                        axis=1
                                    )
                                    st.dataframe(
                                        good_display[['AWC', 'Attendance', 'Attendance %']],
                                        use_container_width=True,
                                        hide_index=True
                                    )
                                    st.info(f"ℹ️ {len(good_awcs)} AWCs with good attendance. Aim for 90%+!")
                                else:
                                    st.info("No AWCs in this category.")

                            with att_tab3:
                                low_awcs = awc_attendance[awc_attendance['Attendance %'] < 75].sort_values('Attendance %')
                                if not low_awcs.empty:
                                    low_display = low_awcs.copy()
                                    low_display['Attendance'] = low_display.apply(
                                        lambda x: f"{int(x['Total Children Attended'])}/{int(x['Total Children 3 yr to 6 yr'])} ({x['Attendance %']:.1f}%)",
                                        axis=1
                                    )
                                    st.dataframe(
                                        low_display[['AWC', 'Attendance', 'Attendance %']],
                                        use_container_width=True,
                                        hide_index=True
                                    )
                                    st.error(f"⚠️ {len(low_awcs)} AWCs need immediate attention!")
                                else:
                                    st.success("✅ No AWCs with low attendance!")
                            # fig_att2 = go.Figure(data=[go.Pie(
                            #     labels=['Excellent (≥90%)', 'Good (75-89%)', 'Needs Improvement (<75%)'],
                            #     values=[excellent, good, needs_improvement],
                            #     hole=0.4,
                            #     marker_colors=['#4CAF50', '#FF9800', '#F44336'],
                            #     textinfo='label+value+percent',
                            #     textfont_size=12
                            # )])
                            fig_att2.update_layout(showlegend=True, height=350, margin=dict(l=20, r=20, t=30, b=20))
                            st.plotly_chart(fig_att2, use_container_width=True)
                        
                        # Show AWCs with low attendance
                        low_attendance = awc_attendance[awc_attendance['Attendance %'] < 75].sort_values('Attendance %')
                        if not low_attendance.empty:
                            st.markdown("---")
                            st.markdown("#### ⚠️ AWCs with Low Attendance (<75%)")
                            
                            low_att_display = low_attendance.copy()
                            low_att_display['Attendance'] = low_att_display.apply(
                                lambda x: f"{int(x['Total Children Attended'])}/{int(x['Total Children 3 yr to 6 yr'])} ({x['Attendance %']:.1f}%)",
                                axis=1
                            )
                            st.dataframe(
                                low_att_display[['AWC', 'Attendance']],
                                use_container_width=True,
                                hide_index=True
                            )
                    
                    st.markdown("---")
                    
                    # AWCs with Issues for this date
                    st.markdown("#### ⚠️ AWCs with Violations on This Date")
                    
                    # Closed AWCs
                    if closed_awcs > 0:
                        st.markdown(f"**🟡 Closed AWCs ({closed_awcs}):**")
                        closed_list = date_data[date_data["AWC didn't open"] == 1]['AWC'].tolist()
                        st.write(", ".join(closed_list))
                        st.markdown("---")
                    
                    # No HCM AWCs
                    if no_hcm_awcs > 0:
                        st.markdown(f"**🟠 AWCs Without HCM ({no_hcm_awcs}):**")
                        no_hcm_list = date_data[
                            (date_data["AWC didn't open"] == 0) & 
                            (date_data["Total HCM Given"] == 0)
                        ]['AWC'].tolist()
                        st.write(", ".join(no_hcm_list))
                        st.markdown("---")
                    
                    # Snack violations
                    if snack_metric > 0:
                        if day_name in ['Tuesday', 'Thursday', 'Saturday']:
                            st.markdown(f"**🔴 AWCs with Missing Snacks ({snack_metric}):**")
                            snack_list = date_data[
                                (date_data["AWC didn't open"] == 0) & 
                                (date_data["Morning Snack Given"] == 0)
                            ]['AWC'].tolist()
                        else:
                            st.markdown(f"**🌸 AWCs with Extra Snacks ({snack_metric}):**")
                            snack_list = date_data[
                                (date_data["AWC didn't open"] == 0) & 
                                (date_data["Morning Snack Given"] != 0)
                            ]['AWC'].tolist()
                        st.write(", ".join(snack_list))
                    
                    # Show detailed data table
                    st.markdown("---")
                    
                    st.markdown("#### 📋 Detailed Records for This Date")
                    
                    # Add visual indicators to the dataframe
                    display_df = date_data.copy()
                    
                    # Create status column
                    def get_status(row):
                        statuses = []
                        if row["AWC didn't open"] == 1:
                            statuses.append("🟡 Closed")
                        elif row["Total HCM Given"] == 0:
                            statuses.append("🟠 No HCM")
                        
                        if row["AWC didn't open"] == 0:
                            if day_name in ['Tuesday', 'Thursday', 'Saturday'] and row["Morning Snack Given"] == 0:
                                statuses.append("🔴 Missing Snack")
                            elif day_name in ['Monday', 'Wednesday', 'Friday'] and row["Morning Snack Given"] != 0:
                                statuses.append("🌸 Extra Snack")
                            
                            # Add attendance status
                            if 'Attendance %' in row:
                                att_pct = row['Attendance %']
                                if att_pct < 75:
                                    statuses.append(f"📉 Low Attendance ({att_pct:.1f}%)")
                        
                        return " | ".join(statuses) if statuses else "✅ Compliant"
                                        
                    display_df['Status'] = display_df.apply(get_status, axis=1)
                    
                    # Reorder columns
                    # cols_to_show = ['AWC', 'Status', "AWC didn't open", "Total HCM Given", "Morning Snack Given"]
                    cols_to_show = ['AWC', 'Status', 'Attendance Marked', 'Attendance %', "AWC didn't open", "Total HCM Given", "Morning Snack Given"]
                    other_cols = [c for c in display_df.columns if c not in cols_to_show and c not in ['Date', 'Day', 'Source_Filename']]
                    final_display_cols = cols_to_show + other_cols
                    
                    st.dataframe(
                        display_df[final_display_cols],
                        use_container_width=True,
                        hide_index=True
                    )
                    
                else:
                    st.warning("No data available for the selected date.")
            else:
                st.warning("No valid dates found in the dataset.")
        
        with tab4:
            st.markdown("#### 🏢 Top AWCs with Issues")
            
            # AWC-wise analysis
            awc_summary = finalized_df.groupby('AWC').agg({
                'Date': 'count',
                "AWC didn't open": 'sum',
                "Total HCM Given": lambda x: (x == 0).sum(),
            }).reset_index()
            awc_summary.columns = ['AWC', 'Total Days', 'Days Closed', 'Days Without HCM']
            awc_summary['Total Issues'] = awc_summary['Days Closed'] + awc_summary['Days Without HCM']
            awc_summary = awc_summary.sort_values('Total Issues', ascending=False)
            
            st.markdown("**Top 20 AWCs with Most Issues**")
            st.dataframe(awc_summary.head(20), use_container_width=True, hide_index=True)
            
            st.markdown("---")
            st.markdown("**Top 10 AWCs by Closure Days**")
            top_closures = awc_summary.nlargest(10, 'Days Closed')[['AWC', 'Days Closed']]
            st.bar_chart(top_closures.set_index('AWC'))
        
        with tab5:
            st.markdown("#### ⚠️ Compliance Issues Breakdown")
            
            # Create issue summary
            issue_data = {
                'Issue Type': [
                    '🟡 AWC Closures',
                    '🟠 No Hot Cooked Meals',
                    '🌸 Extra Snacks (Mon/Wed/Fri)',
                    '🔴 Missing Snacks (Tue/Thu/Sat)'
                ],
                'Count': [yellow_count, orange_count, pink_count, red_count],
                'Percentage': [
                    f"{(yellow_count/len(finalized_df)*100):.1f}%" if len(finalized_df) > 0 else "0%",
                    f"{(orange_count/len(finalized_df)*100):.1f}%" if len(finalized_df) > 0 else "0%",
                    f"{(pink_count/len(finalized_df)*100):.1f}%" if len(finalized_df) > 0 else "0%",
                    f"{(red_count/len(finalized_df)*100):.1f}%" if len(finalized_df) > 0 else "0%"
                ]
            }
            
            issue_df = pd.DataFrame(issue_data)
            st.dataframe(issue_df, use_container_width=True, hide_index=True)
            
            st.markdown("---")
            st.markdown("**Issue Distribution Chart**")
            st.bar_chart(issue_df.set_index('Issue Type')['Count'])
            
            st.markdown("---")
            st.markdown("#### 💡 Recommendations")
            
            if yellow_count > len(finalized_df) * 0.1:
                st.warning(f"⚠️ High closure rate detected ({yellow_count} closures). Consider investigating AWCs with frequent closures.")
            
            if orange_count > len(finalized_df) * 0.05:
                st.warning(f"⚠️ Significant number of days without HCM provision ({orange_count}). Review meal preparation processes.")
            
            if pink_count > 0:
                st.info(f"ℹ️ {pink_count} instances of snacks given on non-snack days. Verify if this follows local guidelines.")
            
            if red_count > len(finalized_df) * 0.05:
                st.error(f"❌ Critical: {red_count} instances of missing snacks on designated snack days. Immediate action required.")
            
            if compliance_rate >= 90:
                st.success("✅ Excellent compliance rate! Keep up the good work.")
            elif compliance_rate >= 70:
                st.info("ℹ️ Good compliance rate, but there's room for improvement.")
            else:
                st.error("❌ Low compliance rate. Urgent attention needed to address issues.")
        
        st.markdown("---")

        # Prepare download
        buffer_name = f"combined_awc_data_final_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        if output_format == "xlsx":
            # STEP 1 — write dataframe normally
            temp_buffer = io.BytesIO()
            
            with pd.ExcelWriter(temp_buffer, engine="openpyxl") as writer:
                finalized_df.to_excel(writer, index=False, sheet_name="combined")

            # STEP 2 — reopen using openpyxl
            temp_buffer.seek(0)
            from openpyxl import load_workbook
            wb = load_workbook(temp_buffer)
            ws = wb["combined"]

            # Build column index mapping
            col_to_idx = {col: idx+1 for idx, col in enumerate(finalized_df.columns)}

            # Create fills
            fills = {
                "FFFF00": PatternFill(start_color="FFFF00", fill_type="solid"),
                "FFA500": PatternFill(start_color="FFA500", fill_type="solid"),
                "FFC0CB": PatternFill(start_color="FFC0CB", fill_type="solid"),
                "FF033E": PatternFill(start_color="FF033E", fill_type="solid"),
            }

            # STEP 3 — Apply color formatting
            for (row_idx, col_name, hex_fill) in formatting_instructions:
                if col_name not in col_to_idx:
                    continue
                excel_row = row_idx + 2
                excel_col = col_to_idx[col_name]
                cell = ws.cell(row=excel_row, column=excel_col)
                if hex_fill in fills:
                    cell.fill = fills[hex_fill]

            # STEP 4 — save final workbook
            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            st.download_button(
                label="📥 Download finalized results (XLSX with colors)",
                data=final_buffer,
                file_name=f"{buffer_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        else:
            csv_bytes = finalized_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Download finalized results (CSV)",
                data=csv_bytes,
                file_name=f"{buffer_name}.csv",
                mime="text/csv"
            )

        st.info("**📋 Summary of Applied Rules:**\n"
                f"- ✓ Rule 0: All Sunday rows and {len(st.session_state.government_holidays)} Government Holiday(s) removed\n"
                "- ✓ Rule 1: 🟡 Yellow = AWC closed (AWC didn't open = 1)\n"
                "- ✓ Rule 2: 🟠 Orange = No hot cooked meals (Total HCM = 0 while open)\n"
                "- ✓ Rule 3: 🌸 Pink = Snacks given on Mon/Wed/Fri (non-snack days)\n"
                "- ✓ Rule 4: 🔴 Red = Missing snacks on Tue/Thu/Sat (snack days)\n\n"
                "📥 Colors are visible in the downloaded XLSX file. Original data values are preserved.")
        
        # Footer Section
        st.markdown("---")
        st.markdown("### 🎨 Understanding the Color Codes in Your Downloaded File")
        
        st.markdown("""
        When you open the downloaded XLSX file, you'll see cells highlighted in four different colors. 
        Here's what each color means:
        """)
        
        footer_col1, footer_col2 = st.columns(2)
        
        with footer_col1:
            st.markdown("""
            <div style="background-color: #FFFF00; padding: 12px; border-radius: 8px; margin-bottom: 10px; border-left: 5px solid #CCCC00;">
                <strong style="color: #000;">🟡 YELLOW - AWC Closure</strong><br>
                <span style="font-size: 14px; color: #333;">
                This cell is highlighted when the AWC was officially closed on that day. 
                The "AWC didn't open" column will show value 1.
                <br><br>
                <strong>What to check:</strong> Verify if the closure was authorized and documented properly.
                </span>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("""
            <div style="background-color: #FFC0CB; padding: 12px; border-radius: 8px; margin-bottom: 10px; border-left: 5px solid #FF90A5;">
                <strong style="color: #000;">🌸 PINK - Unexpected Snack Distribution</strong><br>
                <span style="font-size: 14px; color: #333;">
                This cell appears when snacks were given on Monday, Wednesday, or Friday. 
                These are typically non-snack days in the schedule.
                <br><br>
                <strong>What to check:</strong> Verify if there was a special event or valid reason for snack distribution.
                </span>
            </div>
            """, unsafe_allow_html=True)
        
        with footer_col2:
            st.markdown("""
            <div style="background-color: #FFA500; padding: 12px; border-radius: 8px; margin-bottom: 10px; border-left: 5px solid #CC8400;">
                <strong style="color: #000;">🟠 ORANGE - Missing Hot Cooked Meal</strong><br>
                <span style="font-size: 14px; color: #333;">
                This cell shows when the AWC was open but no Hot Cooked Meal (HCM) was provided. 
                The "Total HCM Given" column will show value 0.
                <br><br>
                <strong>What to check:</strong> This is a compliance issue that requires immediate attention and explanation.
                </span>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("""
            <div style="background-color: #FF033E; padding: 12px; border-radius: 8px; margin-bottom: 10px; border-left: 5px solid #FF97A6;">
                <strong style="color: #000;">🔴 RED - Missing Required Snack</strong><br>
                <span style="font-size: 14px; color: #333;">
                This cell appears when snacks were NOT given on Tuesday, Thursday, or Saturday. 
                These are designated snack days in the weekly schedule.
                <br><br>
                <strong>What to check:</strong> Critical compliance issue. Verify why the scheduled snack was not distributed.
                </span>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        st.markdown("""
        <div style="background-color: #E8F4F8; padding: 20px; border-radius: 10px; border-left: 5px solid #2196F3;">
            <h4 style="margin-top: 0; color: #1976D2;">💡 Quick Action Guide</h4>
            <ul style="color: #333; line-height: 1.8;">
                <li><strong>Yellow cells:</strong> Review closure logs and authorization documents</li>
                <li><strong>Orange cells:</strong> Investigate meal preparation issues and supply chain problems</li>
                <li><strong>Pink cells:</strong> Check for special events or schedule changes that justified extra snacks</li>
                <li><strong>Red cells:</strong> Priority action needed - ensure snack distribution on designated days</li>
            </ul>
            <p style="margin-bottom: 0; color: #555; font-style: italic;">
            💾 Tip: Keep the downloaded XLSX file for your records and use it for compliance reporting.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # App Footer
        st.markdown("""
        <div style="text-align: center; padding: 20px; color: #666;">
            <p style="margin: 5px 0;">🏥 <strong>ICDS AWC Status Tracker</strong></p>
            <p style="margin: 5px 0; font-size: 14px;">Integrated Child Development Services | Anganwadi Centre Monitoring System</p>
            <p style="margin: 5px 0; font-size: 12px;">Track compliance, identify issues, and improve service delivery</p>
        </div>
        """, unsafe_allow_html=True)